# encoding: utf-8

"""
openbikebox Backend
Copyright (C) 2021 binary butterfly GmbH

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <http://www.gnu.org/licenses/>.
"""

import os
from enum import Enum
from uuid import uuid4
from decimal import Decimal
from dataclasses import dataclass
from typing import List, Optional, Any
from datetime import date, datetime, time
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_workbook
from openpyxl.styles import NamedStyle
from flask import render_template
from flask_mail import Message
from webapp.repositories import ActionRepository
from webapp.services.base_service import BaseService
from webapp.common.helpers import unlocalize_datetime, localize_datetime
from webapp.common.mail_helper import MailHelper
from webapp.repositories.action_repository import ExportableAction
from webapp.enum import ActionStatus, PredefinedDaterange


class FieldType(Enum):
    STRING = 'STRING'
    INTEGER = 'INTEGER'
    DECIMAL = 'DECIMAL'
    PRICE = 'PRICE'
    DATE = 'DATE'
    DATETIME = 'DATETIME'


@dataclass
class FieldConfig:

    def __post_init__(self):
        if self.enum_map is None:
            self.enum_map = {}

    title: str
    width: Optional[int] = None
    field_type: Optional[FieldType] = None
    enum_map: Optional[dict] = None


class ActionExporterService(BaseService):

    action_repository: ActionRepository
    mail_helper: MailHelper
    datetime_style = NamedStyle(name='datetime', number_format='dd.mm.yy, hh:mm')

    export_fields = {
        'id': FieldConfig(title='ID', width=6),
        'uid': FieldConfig(title='UID'),
        'source': FieldConfig(title='Quelle'),
        'requested_at': FieldConfig(title='Buchung Start'),
        'paid_at': FieldConfig(title='Buchung Abschluss'),
        'begin': FieldConfig(title='Buchung Beginn'),
        'end': FieldConfig('Buchung Ende'),
        'status': FieldConfig('Status', enum_map={
            ActionStatus.reserved: 'reserviert',
            ActionStatus.booked: 'gebucht',
            ActionStatus.timeouted: 'abgebrochen',
            ActionStatus.cancelled: 'abgebrochen',
            ActionStatus.disrupted: 'unterbrochen',
        }),
        'value_gross': FieldConfig(title='Brutto', width=8, field_type=FieldType.PRICE),
        'value_net': FieldConfig(title='Netto', width=8, field_type=FieldType.PRICE),
        'value_tax': FieldConfig(title='MWSt', width=8, field_type=FieldType.PRICE),
        'user_identifier': FieldConfig(title='User-Referenz'),
        'predefined_daterange': FieldConfig(title='Zeitraum', enum_map={
            PredefinedDaterange.day: 'Tag',
            PredefinedDaterange.week: 'Woche',
            PredefinedDaterange.month: 'Monat',
            PredefinedDaterange.quarter: 'Quartal',
            PredefinedDaterange.year: 'Jahr',
        }),
        'operator_id': FieldConfig(title='Betreiber: ID'),
        'operator_name': FieldConfig(title='Betreiber: Name'),
        'location_id': FieldConfig(title='Ort: ID'),
        'location_name': FieldConfig(title='Ort: Name'),
        'location_address': FieldConfig(title='Ort: Adresse'),
        'location_postalcode': FieldConfig(title='Ort: PLZ'),
        'location_locality': FieldConfig(title='Ort: Stadt'),
        'location_lat': FieldConfig(title='Ort: Breitengrad'),
        'location_lon': FieldConfig(title='Ort: Längengrad'),
        'resource_id': FieldConfig(title='Box: ID'),
        'resource_identifier': FieldConfig(title='Box: UID'),
        'hardware_id': FieldConfig(title='Hardware: ID'),
        'hardware_name': FieldConfig(title='Hardware: Name'),
    }

    def __init__(self, *, action_repository: ActionRepository, mail_helper: MailHelper, **kwargs):
        super().__init__(**kwargs)
        self.action_repository = action_repository
        self.mail_helper = mail_helper

    def export(self, begin: date, end: date) -> str:
        file_name = '%s.xlsx' % uuid4()
        file_path = os.path.join(self.config_helper.get('EXPORT_DIR'), file_name)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Export %s - %s' % (begin.strftime('%d.%m.%Y'), end.strftime('%d.%m.%Y'))
        actions = self.action_repository.fetch_exportable_actions(
            begin=unlocalize_datetime(datetime.combine(begin, time(0))),
            end=unlocalize_datetime(datetime.combine(end, time(0))),
        )
        self.add_header_row(ws=ws)
        row = 2
        for action in actions:
            self.add_row(
                ws=ws,
                row=row,
                action=action
            )
            row += 1
        save_workbook(wb, file_path)
        return file_name

    def add_header_row(self, ws: Worksheet):
        column = 1
        for field, field_config in self.export_fields.items():
            self.add_header_cell(
                ws=ws,
                column=column,
                field=field,
                field_config=field_config
            )
            column += 1

    def add_header_cell(self, ws: Worksheet, column: int, field: str, field_config: FieldConfig):
        ws.cell(column=column, row=1, value=field_config.title)
        ws.column_dimensions[get_column_letter(column)].width = 16 if field_config.width is None else field_config.width

    def add_row(self, ws: Worksheet, row: int, action: ExportableAction):
        column = 1
        for field, field_config in self.export_fields.items():
            self.add_cell(
                ws=ws,
                row=row,
                column=column,
                field=field,
                field_config=field_config,
                value=getattr(action, field)
            )
            column += 1

    def add_cell(self, ws: Worksheet, row: int, column: int, field: str, field_config: FieldConfig, value: Any):
        if value is None:
            return
        if type(value) is datetime:
            ws.cell(row, column, localize_datetime(value).replace(tzinfo=None)).number_format = 'dd.mm.yy, hh:mm'
            return
        if field_config.field_type == FieldType.PRICE:
            ws.cell(row, column, float(value)).number_format = '#,##0.00 €'
            return
        if type(value) is Decimal:
            ws.cell(row, column, float(value))
            return
        if isinstance(value, Enum):
            value = field_config.enum_map.get(value, value.name)
        ws.cell(row, column, value)

    def send_export(self, begin: date, end: date, file_name: str, recipients: List[str]):
        message = Message(
            sender=self.config_helper.get('MAILS_FROM'),
            recipients=recipients,
            subject='Buchungen von %s bis %s' % (begin.strftime('%d.%m.%Y'), end.strftime('%d.%m.%Y')),
            body=render_template(
                'email/action-export.txt',
                project_name=self.config_helper.get('PROJECT_NAME'),
                begin=begin.strftime('%d.%m.%Y'),
                end=end.strftime('%d.%m.%Y'),
            )
        )
        file_path = os.path.join(self.config_helper.get('EXPORT_DIR'), file_name)
        with open(file_path, 'rb') as xlsx_file:
            message.attach(
                filename='%s.xlsx' % date.today().isoformat(),
                content_type='application/vnd.openxmlformats-officedocument. spreadsheetml.sheet',
                data=xlsx_file.read()
            )
        self.mail_helper.send(message)
